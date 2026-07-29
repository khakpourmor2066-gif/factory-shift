import csv
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook


SUPPORTED_SUFFIXES = {".csv", ".xlsx"}


def parse_tabular_file(filename: str, content: bytes) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError("فقط فایل‌های CSV و XLSX پشتیبانی می‌شوند.")
    if not content:
        raise ValueError("فایل خالی است.")
    if suffix == ".csv":
        return _parse_csv(content)
    return _parse_xlsx(content)


def _parse_csv(content: bytes) -> list[dict]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise ValueError("کدگذاری فایل CSV باید UTF-8 باشد.") from error
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise ValueError("سربرگ فایل وجود ندارد.")
    return [_normalize_row(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise ValueError("فایل XLSX معتبر نیست.") from error
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise ValueError("سربرگ فایل وجود ندارد.")
    normalized_headers = [str(value).strip() if value is not None else "" for value in headers]
    if any(not header for header in normalized_headers):
        raise ValueError("سربرگ فایل دارای ستون بدون نام است.")
    return [
        _normalize_row(dict(zip(normalized_headers, values, strict=False)))
        for values in rows
        if any(value is not None and str(value).strip() for value in values)
    ]


def _normalize_row(row: dict) -> dict:
    return {
        str(key).strip(): "" if value is None else str(value).strip()
        for key, value in row.items()
        if key is not None
    }
